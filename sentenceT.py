from openpyxl import load_workbook, Workbook
import numpy as np
import json
import pandas as pd
from scipy.spatial.distance import cdist
import datetime


def load_embeddings_from_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def process_new_descriptions(file_path_takalot_file, embedding_model, processor, ata, reference_json, log_fn=print):
    log_fn("📁 ניקוי של הקובץ אקסל מקורי...")

    # 1. Charger le fichier Excel source
    wb_src = load_workbook(file_path_takalot_file, data_only=True)
    ws_src = wb_src.active

    # 2. Créer un fichier nettoyé contenant uniquement les lignes non vides
    wb_clean = Workbook()
    ws_clean = wb_clean.active

    for row in ws_src.iter_rows():
        has_value = any(cell.value is not None and str(cell.value).strip() != "" for cell in row)
        if has_value:
            for cell in row:
                value = cell.value
                if isinstance(value, datetime.datetime):
                    value = value.date()
                ws_clean.cell(row=cell.row, column=cell.column, value=value)

    wb_clean.save(file_path_takalot_file)

    # 3. Recharger le fichier nettoyé
    wb = load_workbook(file_path_takalot_file)
    ws = wb.active

    # 4. Charger le fichier ATA et créer les deux listes
    ata_df = pd.read_excel(ata)
    ata_df.columns = ata_df.columns.str.strip().str.lower()

    
    
    
    required_cols = {"ata chapter", "categories"}
    if not required_cols.issubset(set(ata_df.columns)):
        raise ValueError(f"❌ Le fichier ATA ne contient pas les colonnes requises : {required_cols}. Colonnes trouvées : {ata_df.columns.tolist()}")

    
    
    
    ata_df["ata chapter"] = ata_df["ata chapter"].astype(str).str.strip().str.lower()
    ata_df["categories"] = ata_df["categories"].str.strip().str.lower()

    ata_general = ata_df.loc[ata_df["categories"] == "general", "ata chapter"].tolist()
    ata_not_cabine = ata_df.loc[ata_df["categories"] == "not cabin", "ata chapter"].tolist()

    # 5. Recherche de la cellule "WO Desc"
    desc_cell = None
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=30):
        for cell in row:
            if cell.value:
                value = str(cell.value).strip().lower()
                if value == "wo desc":
                    desc_cell = cell
        if desc_cell:
            break

    if not desc_cell:
        log_fn("❌ הכותרות לא נמצאו")
        return

    # Nettoyage des lignes au-dessus si besoin
    if desc_cell.row > 1:
        ws.delete_rows(1, desc_cell.row - 1)
        desc_cell = ws.cell(row=1, column=desc_cell.column)

    if ws.cell(row=1, column=1).value is None or str(ws.cell(row=1, column=1).value).strip() == "":
        ws.delete_cols(1)

    # 6. Recherche des colonnes importantes
    main_cell = sub_cell = ata_cell = None
    for cell in ws[desc_cell.row]:
        if cell.value:
            value = str(cell.value).strip().lower()
            if value == "main category":
                main_cell = cell
            elif value == "sub category":
                sub_cell = cell
            elif value == "ata chapter":
                ata_cell = cell

    col_index = desc_cell.column
    start_row = desc_cell.row + 1

    descriptions = []
    row_indexes = []

    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=col_index).value
        if value is not None and str(value).strip():
            descriptions.append(str(value))
            row_indexes.append(row)
            log_fn(f"שורות שנקראו : {row}")
        else:
            break

    log_fn("🔍 ניקוי של התיאורים עבור כל תקלה...")
    cleaned = [processor.clean_text(text) for text in descriptions]

    log_fn("📦 טעינה של וקטורי ייחוס...")
    reference_embeddings = load_embeddings_from_json(reference_json)
    ref_keys = list(reference_embeddings.keys())
    ref_matrix = np.array([reference_embeddings[k] for k in ref_keys])

    # 7. Préparation des lignes à classifier par vecteurs
    log_fn("🧠 מיון לפי ATA והכנת תיאורים שאינם ידועים...")
    rows_to_classify = []
    descriptions_to_classify = []
    index_in_cleaned = []

    for i, row_num in enumerate(row_indexes):
        ata_value = ws.cell(row=row_num, column=ata_cell.column).value
        ata_value = str(ata_value).strip().lower() if ata_value else ""

        if ata_value in ata_general:
            ws.cell(row=row_num, column=main_cell.column).value = "General"
            ws.cell(row=row_num, column=sub_cell.column).value = "General"
        elif ata_value in ata_not_cabine:
            ws.cell(row=row_num, column=main_cell.column).value = "Not Cabin"
            ws.cell(row=row_num, column=sub_cell.column).value = "Not Cabin"
        else:
            rows_to_classify.append(row_num)
            descriptions_to_classify.append(cleaned[i])
            index_in_cleaned.append(i)
            

    # 8. Traitement des descriptions restantes
    if descriptions_to_classify:
        log_fn(f"🧪 סיווג {len(descriptions_to_classify)} תקלות שלא הופיעו ברשימות ATA מתוך {len(descriptions)}")

        takalot_embeddings = embedding_model.encode(
            descriptions_to_classify,
            normalize_embeddings=True,
            show_progress_bar=True,
            batch_size=32
        )

        dist_matrix = cdist(takalot_embeddings, ref_matrix, metric="euclidean")
        top_1_idx = np.argmin(dist_matrix, axis=1)

        for i, row_num in enumerate(rows_to_classify):
            best = ref_keys[top_1_idx[i]]
            cat_1, subcat_1 = best.split(" | ")
            ws.cell(row=row_num, column=main_cell.column).value = cat_1
            ws.cell(row=row_num, column=sub_cell.column).value = subcat_1

    wb.save(file_path_takalot_file)
