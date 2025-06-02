import openpyxl
import pandas as pd
import json
from sentence_transformers import SentenceTransformer
from config import file_path_categories, file_path_json, MODEL_DIR_1


class CategoryEmbeddingBuilder:
    def __init__(self, file_path, model_dir):
        self.file_path = file_path
        self.model = SentenceTransformer(model_dir)
        self.wb = openpyxl.load_workbook(file_path, data_only=True)

        self.main_df = self.load_main_categories()
        self.sub_df = self.load_sub_categories()
        self.merged_df = self.merge_dataframes()

    def load_main_categories(self):
        ws = self.wb["Main_Categories"]
        data = [(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] and row[1]]
        return pd.DataFrame(data, columns=["MAIN_CATEGORY", "MAIN_DEFINITION"])

    def load_sub_categories(self):
        ws = self.wb["Sub_Categories"]
        data = [(row[0], row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] and row[1] and row[2]]
        return pd.DataFrame(data, columns=["MAIN_CATEGORY", "SUB_CATEGORY", "SUB_DEFINITION"])

    def merge_dataframes(self):
        df = pd.merge(self.sub_df, self.main_df, on="MAIN_CATEGORY", how="left")
        return df.fillna('')

    def generate_embeddings(self):
        embeddings_dict = {}

        for _, row in self.merged_df.iterrows():
            key = f"{row['MAIN_CATEGORY']} | {row['SUB_CATEGORY']}"
            full_text = f"{row['MAIN_CATEGORY']}, {row['MAIN_DEFINITION']}, {row['SUB_CATEGORY']}, {row['SUB_DEFINITION']}"
            embedding = self.model.encode(full_text, normalize_embeddings=True).tolist()
            embeddings_dict[key] = embedding

        return embeddings_dict

    def save_outputs(self, json_path = file_path_json):
        embeddings_dict = self.generate_embeddings()

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(embeddings_dict, f, indent=4)
        print(f"✅ JSON created: {json_path}")


if __name__ == "__main__":
    builder = CategoryEmbeddingBuilder(file_path_categories, MODEL_DIR_1)
    builder.save_outputs()
    print("c'est fait mon cochon")
